"""
Export and import profiles via Excel.
"""

import io
from typing import Any

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .models import BrowserSettings, Profile, ProfileStatus, WebRTCMode
from .profile_manager import ProfileManager


class ExcelManager:
    """Handle Excel export and import of profiles."""

    def __init__(self, profile_manager: ProfileManager):
        self.profile_manager = profile_manager

        # Columns for export: (field, header, help text)
        self.columns = [
            ("id", "Profile ID", "Auto-generated ID (read-only)"),
            ("name", "Profile name", "Unique profile name"),
            ("group", "Group", "Profile group name"),
            ("status", "Status", "active or inactive"),
            ("os", "Operating system", "windows, macos or linux"),
            ("screen", "Screen resolution", "For example: 1920x1080"),
            ("window_width", "Window width", "Browser window width (800-3840)"),
            ("window_height", "Window height", "Browser window height (600-2160)"),
            ("languages", "Browser languages", "Comma-separated: en-US, en"),
            ("timezone", "Timezone", "For example: Europe/Moscow"),
            ("locale", "Locale", "For example: ru_RU"),
            ("webrtc_mode", "WebRTC mode", "forward, replace, real or none"),
            ("canvas_noise", "Canvas noise", "true or false"),
            ("webgl_noise", "WebGL noise", "true or false"),
            ("audio_noise", "Audio noise", "true or false"),
            ("stable_canvas", "Stable canvas", "true or false; same canvas every launch"),
            ("hardware_concurrency", "CPU cores", "Number of cores (1-32)"),
            ("device_memory", "Device memory", "Memory in GB (1-128)"),
            ("max_touch_points", "Touch points", "Number of touch points (0-10)"),
            ("proxy_type", "Proxy type", "http, https, socks4, socks5 or empty"),
            ("proxy_server", "Proxy server", "host:port or empty"),
            ("proxy_username", "Proxy username", "Username or empty"),
            ("proxy_password", "Proxy password", "Password or empty"),
            ("geo_mode", "Geolocation mode", "auto or manual"),
            ("geo_latitude", "Latitude", "Latitude coordinate (-90 to 90)"),
            ("geo_longitude", "Longitude", "Longitude coordinate (-180 to 180)"),
            ("geo_accuracy", "Geolocation accuracy", "Accuracy in meters (1-10000)"),
            ("notes", "Notes", "Additional information about the profile"),
        ]

    async def export_profiles_to_excel(self, profiles: list[Profile] | None = None) -> bytes:
        """Export profiles to an Excel file."""
        logger.info("Exporting profiles to Excel")

        if profiles is None:
            profiles = await self.profile_manager.list_profiles()

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Browser profiles"

        # Configure styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Add headers
        for col_idx, (_field, header, comment_text) in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

            # Add an explanatory comment
            comment = Comment(comment_text, "CamoufoxProfileManager")
            cell.comment = comment

        # Add profile rows
        for row_idx, profile in enumerate(profiles, 2):
            self._write_profile_row(ws, row_idx, profile)

        # Add data validation
        self._add_data_validation(ws, len(profiles) + 10)  # +10 rows for new profiles

        # Auto-size column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Serialize to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(f"Export complete: {len(profiles)} profiles")
        return excel_buffer.getvalue()

    def _write_profile_row(self, ws, row_idx: int, profile: Profile):
        """Write a profile as a row in the worksheet."""
        browser_settings = profile.browser_settings
        proxy = profile.proxy

        # Prepare the row data
        row_data = {
            "id": profile.id,
            "name": profile.name,
            "group": profile.group or "",
            "status": profile.status.value
            if hasattr(profile.status, "value")
            else str(profile.status),
            "os": browser_settings.os,
            "screen": browser_settings.screen,
            "window_width": browser_settings.window_width or "",
            "window_height": browser_settings.window_height or "",
            "languages": ", ".join(browser_settings.languages),
            "timezone": browser_settings.timezone or "",
            "locale": browser_settings.locale or "",
            "webrtc_mode": browser_settings.webrtc_mode.value
            if hasattr(browser_settings.webrtc_mode, "value")
            else str(browser_settings.webrtc_mode),
            "canvas_noise": str(browser_settings.canvas_noise).lower(),
            "webgl_noise": str(browser_settings.webgl_noise).lower(),
            "audio_noise": str(browser_settings.audio_noise).lower(),
            "stable_canvas": str(browser_settings.stable_canvas).lower(),
            "hardware_concurrency": browser_settings.hardware_concurrency or "",
            "device_memory": browser_settings.device_memory or "",
            "max_touch_points": browser_settings.max_touch_points,
            "proxy_type": proxy.type.value
            if proxy and hasattr(proxy.type, "value")
            else (str(proxy.type) if proxy else ""),
            "proxy_server": proxy.server if proxy else "",
            "proxy_username": proxy.username if proxy else "",
            "proxy_password": proxy.password if proxy else "",
            "geo_mode": "manual" if browser_settings.geolocation else "auto",
            "geo_latitude": browser_settings.geolocation.get("lat", "")
            if browser_settings.geolocation
            else "",
            "geo_longitude": browser_settings.geolocation.get("lon", "")
            if browser_settings.geolocation
            else "",
            "geo_accuracy": browser_settings.geolocation.get("accuracy", "")
            if browser_settings.geolocation
            else "",
            "notes": profile.notes or "",
        }

        # Write the values into the row
        for col_idx, (field, _, _) in enumerate(self.columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))

    def _add_data_validation(self, ws, max_row: int):
        """Add data validation for the columns."""
        # Status validation
        status_validation = DataValidation(
            type="list", formula1='"active,inactive"', allow_blank=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"D2:D{max_row}")

        # OS validation
        os_validation = DataValidation(
            type="list", formula1='"windows,macos,linux"', allow_blank=True
        )
        ws.add_data_validation(os_validation)
        os_validation.add(f"E2:E{max_row}")

        # WebRTC mode validation
        webrtc_validation = DataValidation(
            type="list", formula1='"forward,replace,real,none"', allow_blank=True
        )
        ws.add_data_validation(webrtc_validation)
        webrtc_validation.add(f"L2:L{max_row}")

        # Boolean field validation
        bool_validation = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"M2:O{max_row}")  # Canvas, WebGL, Audio noise

        # Proxy type validation
        proxy_validation = DataValidation(
            type="list", formula1='"http,https,socks4,socks5"', allow_blank=True
        )
        ws.add_data_validation(proxy_validation)
        proxy_validation.add(f"S2:S{max_row}")

        # Geolocation mode validation
        geo_validation = DataValidation(type="list", formula1='"auto,manual"', allow_blank=True)
        ws.add_data_validation(geo_validation)
        geo_validation.add(f"W2:W{max_row}")

    async def import_profiles_from_excel(self, excel_data: bytes) -> dict[str, Any]:
        """Import profiles from an Excel file."""
        logger.info("Importing profiles from Excel")

        result: dict[str, Any] = {
            "success": True,
            "created_count": 0,
            "updated_count": 0,  # Always 0: import only creates new profiles
            "error_count": 0,
            "errors": [],
            "summary": "",
        }

        try:
            # Load the workbook
            excel_buffer = io.BytesIO(excel_data)
            wb = load_workbook(excel_buffer)
            ws = wb.active

            # Read the headers
            headers = {}
            for col_idx, (field, _, _) in enumerate(self.columns, 1):
                headers[field] = col_idx

            # Process the rows
            for row_idx in range(2, ws.max_row + 1):
                try:
                    await self._process_excel_row(ws, row_idx, headers, result)
                except Exception as e:
                    error_msg = f"Row {row_idx}: {e}"
                    result["errors"].append(error_msg)
                    result["error_count"] += 1
                    logger.error(error_msg)

            # Build the summary report
            result["summary"] = (
                f"Import complete:\n"
                f"Profiles created: {result['created_count']}\n"
                f"Errors: {result['error_count']}\n"
                f"All profiles are created with new auto-generated IDs"
            )

            if result["error_count"] == 0:
                logger.info(f"Import succeeded: created {result['created_count']} profiles")
            else:
                logger.warning(f"Import finished with {result['error_count']} errors")
                result["success"] = False

        except Exception as e:
            logger.error(f"Fatal import error: {e}")
            result["success"] = False
            result["errors"].append(f"Fatal error: {e}")
            result["summary"] = "Import failed due to a fatal error"

        return result

    async def _process_excel_row(
        self, ws, row_idx: int, headers: dict[str, int], result: dict[str, Any]
    ):
        """Process a single Excel row."""
        # Read the values from the row
        row_data = {}
        for field, col_idx in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[field] = str(cell_value).strip() if cell_value is not None else ""

        # Skip empty rows
        if not any(row_data.values()):
            return

        # Validate required fields
        if not row_data["name"]:
            raise ValueError("Profile name is required")

        # The ID is ignored: a new profile with an auto-generated ID is always created
        # This keeps IDs unique and sequential
        browser_settings = self._create_browser_settings(row_data)
        proxy_config = self._create_proxy_config(row_data)

        await self.profile_manager.create_profile(
            name=row_data["name"],
            group=row_data["group"] or None,
            browser_settings=browser_settings,
            proxy_config=proxy_config,
            generate_fingerprint=False,  # Use the data from Excel
        )
        result["created_count"] += 1

    def _prepare_profile_updates(self, row_data: dict[str, str]) -> dict[str, Any]:
        """Build the profile-update payload from a row."""
        updates: dict[str, Any] = {}

        # Core profile fields
        if row_data["name"]:
            updates["name"] = row_data["name"]
        if row_data["group"]:
            updates["group"] = row_data["group"]
        if row_data["status"]:
            updates["status"] = ProfileStatus(row_data["status"])
        if row_data["notes"]:
            updates["notes"] = row_data["notes"]

        # Browser settings
        browser_updates: dict[str, Any] = {}
        if row_data["os"]:
            browser_updates["os"] = row_data["os"]
        if row_data["screen"]:
            browser_updates["screen"] = row_data["screen"]
        if row_data["window_width"]:
            browser_updates["window_width"] = int(row_data["window_width"])
        if row_data["window_height"]:
            browser_updates["window_height"] = int(row_data["window_height"])
        if row_data["languages"]:
            browser_updates["languages"] = [
                lang.strip() for lang in row_data["languages"].split(",")
            ]
        if row_data["timezone"]:
            browser_updates["timezone"] = row_data["timezone"]
        if row_data["locale"]:
            browser_updates["locale"] = row_data["locale"]
        if row_data["webrtc_mode"]:
            browser_updates["webrtc_mode"] = WebRTCMode(row_data["webrtc_mode"])
        if row_data["canvas_noise"]:
            browser_updates["canvas_noise"] = row_data["canvas_noise"].lower() == "true"
        if row_data["webgl_noise"]:
            browser_updates["webgl_noise"] = row_data["webgl_noise"].lower() == "true"
        if row_data["audio_noise"]:
            browser_updates["audio_noise"] = row_data["audio_noise"].lower() == "true"
        if row_data["stable_canvas"]:
            browser_updates["stable_canvas"] = row_data["stable_canvas"].lower() == "true"
        if row_data["hardware_concurrency"]:
            browser_updates["hardware_concurrency"] = int(row_data["hardware_concurrency"])
        if row_data["device_memory"]:
            browser_updates["device_memory"] = int(row_data["device_memory"])
        if row_data["max_touch_points"]:
            browser_updates["max_touch_points"] = int(row_data["max_touch_points"])

        # Geolocation
        if (
            row_data["geo_mode"] == "manual"
            and row_data["geo_latitude"]
            and row_data["geo_longitude"]
        ):
            browser_updates["geolocation"] = {
                "lat": float(row_data["geo_latitude"]),
                "lon": float(row_data["geo_longitude"]),
                "accuracy": int(row_data["geo_accuracy"]) if row_data["geo_accuracy"] else 10,
            }
        elif row_data["geo_mode"] == "auto":
            browser_updates["geolocation"] = None

        if browser_updates:
            updates["browser_settings"] = browser_updates

        # Proxy
        if row_data["proxy_type"] and row_data["proxy_server"]:
            updates["proxy_config"] = {
                "type": row_data["proxy_type"],
                "server": row_data["proxy_server"],
                "username": row_data["proxy_username"] or None,
                "password": row_data["proxy_password"] or None,
            }
        elif not row_data["proxy_type"]:
            updates["proxy_config"] = None

        return updates

    def _create_browser_settings(self, row_data: dict[str, str]) -> BrowserSettings:
        """Build a BrowserSettings object from Excel data."""
        # Geolocation
        geolocation = None
        if (
            row_data["geo_mode"] == "manual"
            and row_data["geo_latitude"]
            and row_data["geo_longitude"]
        ):
            geolocation = {
                "lat": float(row_data["geo_latitude"]),
                "lon": float(row_data["geo_longitude"]),
                "accuracy": int(row_data["geo_accuracy"]) if row_data["geo_accuracy"] else 10,
            }

        return BrowserSettings(
            os=row_data["os"] or "windows",
            screen=row_data["screen"] or "1920x1080",
            window_width=int(row_data["window_width"]) if row_data["window_width"] else 1280,
            window_height=int(row_data["window_height"]) if row_data["window_height"] else 720,
            languages=[lang.strip() for lang in row_data["languages"].split(",")]
            if row_data["languages"]
            else ["en-US", "en"],
            timezone=row_data["timezone"] or "UTC",
            locale=row_data["locale"] or "en_US",
            webrtc_mode=WebRTCMode(row_data["webrtc_mode"])
            if row_data["webrtc_mode"]
            else WebRTCMode.REPLACE,
            canvas_noise=row_data["canvas_noise"].lower() == "true"
            if row_data["canvas_noise"]
            else True,
            webgl_noise=row_data["webgl_noise"].lower() == "true"
            if row_data["webgl_noise"]
            else True,
            audio_noise=row_data["audio_noise"].lower() == "true"
            if row_data["audio_noise"]
            else True,
            stable_canvas=row_data["stable_canvas"].lower() == "true"
            if row_data["stable_canvas"]
            else False,
            hardware_concurrency=int(row_data["hardware_concurrency"])
            if row_data["hardware_concurrency"]
            else 4,
            device_memory=int(row_data["device_memory"]) if row_data["device_memory"] else 8,
            max_touch_points=int(row_data["max_touch_points"])
            if row_data["max_touch_points"]
            else 0,
            geolocation=geolocation,
        )

    def _create_proxy_config(self, row_data: dict[str, str]) -> dict[str, Any] | None:
        """Build a proxy configuration from Excel data."""
        if not row_data["proxy_type"] or not row_data["proxy_server"]:
            return None

        return {
            "type": row_data["proxy_type"],
            "server": row_data["proxy_server"],
            "username": row_data["proxy_username"] or None,
            "password": row_data["proxy_password"] or None,
        }
