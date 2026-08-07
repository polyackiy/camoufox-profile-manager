"""
Interactive wizard for migrating Chrome profiles into Camoufox.
"""

import asyncio
import sys
from pathlib import Path

from loguru import logger

# Add the module path
sys.path.append(str(Path(__file__).parent))

from camoufox_pm.core.database import StorageManager
from camoufox_pm.core.profile_manager import ProfileManager

from .migration_manager import ChromeMigrationManager


async def run_chrome_migration_wizard():
    """Interactive Chrome profile migration wizard."""

    logger.info("Starting the Chrome profile migration wizard")

    # Initialize components
    storage = StorageManager()
    await storage.initialize()

    profile_manager = ProfileManager(storage)
    await profile_manager.initialize()

    migration_manager = ChromeMigrationManager(profile_manager)

    try:
        # 1. Discover Chrome profiles
        logger.info("Step 1: Discover Chrome profiles")
        chrome_profiles = await migration_manager.discover_chrome_profiles()

        if not chrome_profiles:
            logger.warning("No Chrome profiles found")
            logger.info("Make sure Chrome is installed and has profiles with data")

            # Show the search paths
            paths = migration_manager.chrome_importer.chrome_data_paths
            logger.info(f"Searched in: {paths}")
            return

        logger.info(f"Found {len(chrome_profiles)} Chrome profiles:")
        for i, profile in enumerate(chrome_profiles, 1):
            logger.info(f"  {i}. {profile['display_name']} ({profile['name']})")
            logger.info(f"     Path: {profile['path']}")
            logger.info(f"     Cookies: {'yes' if profile['has_cookies'] else 'no'}")
            logger.info(f"     Suggestion: {profile['suggested_mapping']['type']}")

        # 2. Generate the mapping template
        logger.info("\nStep 2: Generate the mapping template")
        template_path = await migration_manager.generate_mapping_template()
        logger.info(f"Template created: {template_path}")

        # 3. Migration status
        logger.info("\nStep 3: Current migration status")
        status = await migration_manager.get_migration_status()
        logger.info(f"Chrome profiles: {status['chrome_profiles_found']}")
        logger.info(f"Total Camoufox profiles: {status['camoufox_profiles_total']}")
        logger.info(f"Already migrated: {status['migrated_profiles']}")

        # 4. Migration dry run
        logger.info("\nStep 4: Migration dry run")
        dry_run_results = await migration_manager.migrate_all_profiles(dry_run=True)

        logger.info(f"Profiles found for migration: {dry_run_results['chrome_profiles_found']}")
        for result in dry_run_results["migration_results"]:
            logger.info(
                f"  - {result['chrome_display_name']}: {result['suggested_mapping']['type']}"
            )

        # 5. Interactive migration with profile selection
        logger.info("\nStep 5: Interactive migration")

        if chrome_profiles:
            print("\nChrome profiles available for migration:")
            for i, profile in enumerate(chrome_profiles, 1):
                status = "✓" if profile["has_cookies"] else "✗"
                print(f"  {i}. {profile['display_name']} ({profile['name']}) - cookies: {status}")

            print("  0. Skip migration")

            try:
                choice = input(
                    f"\nSelect a profile to migrate (1-{len(chrome_profiles)}, 0 to skip): "
                ).strip()
                choice_num = int(choice)

                if choice_num == 0:
                    logger.info("Migration skipped")
                elif 1 <= choice_num <= len(chrome_profiles):
                    selected_profile = chrome_profiles[choice_num - 1]

                    logger.info(
                        f"Starting migration of profile: {selected_profile['display_name']}"
                    )
                    logger.info(f"   Path: {selected_profile['path']}")
                    logger.info(f"   Cookies: {'yes' if selected_profile['has_cookies'] else 'no'}")

                    # Ask for extra parameters
                    print(f"\nMigration settings for profile '{selected_profile['display_name']}':")

                    # New profile name
                    if (
                        selected_profile["display_name"]
                        and selected_profile["display_name"] != selected_profile["name"]
                    ):
                        default_name = f"Chrome - {selected_profile['display_name']} ({selected_profile['name']})"
                    else:
                        default_name = f"Chrome - {selected_profile['name']}"
                    new_name = input(f"New profile name [{default_name}]: ").strip()
                    if not new_name:
                        new_name = default_name

                    # Profile group
                    default_group = "chrome_imports"
                    new_group = input(f"Profile group [{default_group}]: ").strip()
                    if not new_group:
                        new_group = default_group

                    # What to migrate
                    include_cookies = input("Migrate cookies? [Y/n]: ").strip().lower() not in [
                        "n",
                        "no",
                    ]
                    include_bookmarks = input("Migrate bookmarks? [Y/n]: ").strip().lower() not in [
                        "n",
                        "no",
                    ]
                    include_history = input("Migrate history? [y/N]: ").strip().lower() in [
                        "y",
                        "yes",
                    ]

                    # Build the migration mapping
                    mapping = {
                        "create_new_profile": True,
                        "new_profile_name": new_name,
                        "new_profile_group": new_group,
                        "migration_settings": {
                            "include_cookies": include_cookies,
                            "include_bookmarks": include_bookmarks,
                            "include_history": include_history,
                        },
                    }

                    logger.info("\nStarting migration with settings:")
                    logger.info(f"   New name: {new_name}")
                    logger.info(f"   Group: {new_group}")
                    logger.info(f"   Cookies: {'yes' if include_cookies else 'no'}")
                    logger.info(f"   Bookmarks: {'yes' if include_bookmarks else 'no'}")
                    logger.info(f"   History: {'yes' if include_history else 'no'}")

                    result = await migration_manager.migrate_profile(selected_profile, mapping)

                    if result["success"]:
                        logger.info("Migration succeeded!")
                        logger.info(
                            f"   Created profile: {result['camoufox_profile_name']} ({result['camoufox_profile_id']})"
                        )
                        logger.info(
                            f"   Cookies: {result['migration_details'].get('cookies_imported', 0)}"
                        )
                        logger.info(
                            f"   Bookmarks: {result['migration_details'].get('bookmarks_imported', 0)}"
                        )
                        logger.info(
                            f"   History: {result['migration_details'].get('history_imported', 0)}"
                        )
                    else:
                        logger.error("Migration failed:")
                        for error in result["errors"]:
                            logger.error(f"   - {error}")
                else:
                    logger.warning("Invalid profile choice")
                    logger.info("Migration skipped")

            except ValueError:
                logger.warning("Invalid input; a number is expected.")
                logger.info("Migration skipped")
            except KeyboardInterrupt:
                logger.info("\nMigration cancelled by the user")

        # 6. Show the final status
        logger.info("\nStep 6: Final status")
        final_status = await migration_manager.get_migration_status()
        logger.info(f"Migrated profiles: {final_status['migrated_profiles']}")

        if final_status["migrated_profile_details"]:
            logger.info("Migrated profiles:")
            for profile in final_status["migrated_profile_details"]:
                logger.info(
                    f"  - {profile['name']} (ID: {profile['id']}, group: {profile['group']})"
                )

    except Exception as e:
        logger.error(f"Migration error: {e}")
        import traceback

        logger.error(traceback.format_exc())

    finally:
        logger.info("Migration wizard finished")


async def show_excel_integration():
    """Create an example Excel integration for profile mapping."""

    logger.info("Creating the example Excel integration")

    # Create an example Excel mapping file
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Chrome Migration Mapping"

        # Headers
        headers = [
            "Chrome Profile",
            "Chrome Display Name",
            "Chrome Path",
            "Action",
            "Camoufox Profile ID",
            "New Profile Name",
            "New Profile Group",
            "Include Cookies",
            "Include Bookmarks",
            "Include History",
            "Notes",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Example data
        example_data = [
            [
                "Default",
                "Main",
                "/path/to/default",
                "create_new",
                "",
                "Chrome - Main",
                "chrome_imports",
                "TRUE",
                "TRUE",
                "FALSE",
                "Main user profile",
            ],
            [
                "Profile 1",
                "Work",
                "/path/to/profile1",
                "create_new",
                "",
                "Chrome - Work",
                "work_profiles",
                "TRUE",
                "TRUE",
                "FALSE",
                "Work profile",
            ],
            [
                "Profile 2",
                "Personal",
                "/path/to/profile2",
                "use_existing",
                "xcj2cs4r",
                "",
                "",
                "TRUE",
                "FALSE",
                "FALSE",
                "Migrate into an existing profile",
            ],
        ]

        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Save the file
        excel_path = "chrome_migration_example.xlsx"
        wb.save(excel_path)

        logger.info(f"Created the example Excel file: {excel_path}")
        logger.info("This file can be edited and used for bulk migration")

    except ImportError:
        logger.warning("openpyxl is not installed; skipping the Excel example")
    except Exception as e:
        logger.error(f"Failed to create the Excel file: {e}")


def show_chrome_paths():
    """Show the Chrome data paths for the current OS."""
    from .importer import ChromeProfileImporter

    importer = ChromeProfileImporter()

    logger.info("Chrome path information:")
    logger.info(f"Operating system: {importer.system}")
    logger.info(f"Profiles path: {importer.chrome_data_paths['profiles']}")

    # Check that the path exists
    chrome_path = Path(importer.chrome_data_paths["profiles"])
    if chrome_path.exists():
        logger.info("Chrome path found")

        # Show the contents
        try:
            subdirs = [item.name for item in chrome_path.iterdir() if item.is_dir()]
            logger.info(f"Found directories: {subdirs}")
        except Exception as e:
            logger.warning(f"Failed to read the contents: {e}")
    else:
        logger.warning("Chrome path not found")
        logger.info("Possible reasons:")
        logger.info("   - Chrome is not installed")
        logger.info("   - Chrome data is stored elsewhere")
        logger.info("   - No access to the directory")


async def run_bulk_migration_wizard():
    """Bulk Chrome profile migration wizard."""
    logger.info("Starting the bulk Chrome profile migration wizard")

    try:
        # Initialization
        storage_manager = StorageManager("data/profiles.db")
        await storage_manager.initialize()

        profile_manager = ProfileManager(storage_manager, "data")
        await profile_manager.initialize()

        migration_manager = ChromeMigrationManager(
            profile_manager=profile_manager, config_path="config/chrome_migration_config.yaml"
        )

        # Discover profiles
        chrome_profiles = await migration_manager.discover_chrome_profiles()

        if not chrome_profiles:
            logger.warning("No Chrome profiles found")
            return

        logger.info(f"Found {len(chrome_profiles)} Chrome profiles")

        print("\nAvailable Chrome profiles:")
        for i, profile in enumerate(chrome_profiles, 1):
            status = "✓" if profile["has_cookies"] else "✗"
            print(f"  {i}. {profile['display_name']} ({profile['name']}) - cookies: {status}")

        print("\nSelect profiles to migrate:")
        print("   Enter numbers separated by commas (e.g. 1,3,5)")
        print("   Or 'all' to migrate every profile")
        print("   Or '0' to cancel")

        try:
            choice = input("\nYour choice: ").strip()

            if choice == "0":
                logger.info("Bulk migration cancelled")
                return

            selected_profiles = []

            if choice.lower() == "all":
                selected_profiles = chrome_profiles
                logger.info(f"Selected all {len(chrome_profiles)} profiles")
            else:
                # Parse the profile numbers
                profile_numbers = [int(x.strip()) for x in choice.split(",")]
                for num in profile_numbers:
                    if 1 <= num <= len(chrome_profiles):
                        selected_profiles.append(chrome_profiles[num - 1])
                    else:
                        logger.warning(f"Invalid profile number: {num}")

                if not selected_profiles:
                    logger.warning("No valid profile selected")
                    return

                logger.info(f"Selected {len(selected_profiles)} profiles:")
                for profile in selected_profiles:
                    logger.info(f"   - {profile['display_name']}")

            # Shared migration settings
            print(f"\nShared settings for {len(selected_profiles)} profiles:")

            default_group = "chrome_bulk_import"
            new_group = input(f"Group for all profiles [{default_group}]: ").strip()
            if not new_group:
                new_group = default_group

            include_cookies = input("Migrate cookies? [Y/n]: ").strip().lower() not in ["n", "no"]
            include_bookmarks = input("Migrate bookmarks? [Y/n]: ").strip().lower() not in [
                "n",
                "no",
            ]
            include_history = input("Migrate history? [y/N]: ").strip().lower() in ["y", "yes"]

            logger.info(f"\nStarting bulk migration of {len(selected_profiles)} profiles...")

            successful_migrations = 0
            failed_migrations = 0

            for i, profile in enumerate(selected_profiles, 1):
                logger.info(
                    f"\n[{i}/{len(selected_profiles)}] Migrating: {profile['display_name']}"
                )

                # Build a unique profile name
                if profile["display_name"] and profile["display_name"] != profile["name"]:
                    # If the display name differs from the directory name
                    unique_name = f"Chrome - {profile['display_name']} ({profile['name']})"
                else:
                    # If the display name is missing or equals the directory name
                    unique_name = f"Chrome - {profile['name']}"

                mapping = {
                    "create_new_profile": True,
                    "new_profile_name": unique_name,
                    "new_profile_group": new_group,
                    "migration_settings": {
                        "include_cookies": include_cookies,
                        "include_bookmarks": include_bookmarks,
                        "include_history": include_history,
                    },
                }

                try:
                    result = await migration_manager.migrate_profile(profile, mapping)

                    if result["success"]:
                        successful_migrations += 1
                        logger.info(
                            f"[{i}/{len(selected_profiles)}] {profile['display_name']} - succeeded!"
                        )
                        logger.info(
                            f"   Profile: {result['camoufox_profile_name']} ({result['camoufox_profile_id']})"
                        )
                    else:
                        failed_migrations += 1
                        logger.error(
                            f"[{i}/{len(selected_profiles)}] {profile['display_name']} - failed!"
                        )
                        for error in result["errors"]:
                            logger.error(f"   - {error}")

                except Exception as e:
                    failed_migrations += 1
                    logger.error(
                        f"[{i}/{len(selected_profiles)}] {profile['display_name']} - exception: {e}"
                    )

            # Totals
            logger.info("\nBulk migration results:")
            logger.info(f"   Total profiles: {len(selected_profiles)}")
            logger.info(f"   Succeeded: {successful_migrations}")
            if failed_migrations > 0:
                logger.error(f"   Failed: {failed_migrations}")
            else:
                logger.info(f"   Failed: {failed_migrations}")

        except ValueError as e:
            logger.warning(f"Input error: {e}")
        except KeyboardInterrupt:
            logger.info("\nBulk migration cancelled by the user")

    except Exception as e:
        logger.error(f"Bulk migration error: {e}")
        import traceback

        logger.error(traceback.format_exc())

    finally:
        logger.info("Bulk migration wizard finished")


def main():
    """Entry point."""
    logger.info("🦊 Camoufox Chrome Migration Wizard")
    logger.info("=" * 50)

    # Show the Chrome paths
    show_chrome_paths()

    print("\nChoose an action:")
    print("1. Interactive single-profile migration")
    print("2. Bulk migration of several profiles")
    print("3. Create an example Excel configuration")
    print("4. Show Chrome path information")
    print("0. Exit")

    choice = input("\nYour choice (1-4, 0 to exit): ").strip()

    if choice == "1":
        asyncio.run(run_chrome_migration_wizard())
    elif choice == "2":
        asyncio.run(run_bulk_migration_wizard())
    elif choice == "3":
        asyncio.run(show_excel_integration())
    elif choice == "4":
        show_chrome_paths()
    elif choice == "0":
        logger.info("Goodbye!")
    else:
        logger.warning("Invalid choice")


if __name__ == "__main__":
    main()
