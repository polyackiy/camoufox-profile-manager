"""
Интерактивный мастер для миграции профилей Chrome в Camoufox
"""

import asyncio
import sys
from pathlib import Path
from loguru import logger

# Добавляем путь к модулям
sys.path.append(str(Path(__file__).parent))

from core.database import StorageManager
from core.profile_manager import ProfileManager
from core.chrome_migration_manager import ChromeMigrationManager


async def run_chrome_migration_wizard():
    """Интерактивный мастер миграции профилей Chrome"""
    
    logger.info("🚀 Запуск мастера миграции Chrome профилей")
    
    # Инициализация компонентов
    storage = StorageManager()
    await storage.initialize()
    
    profile_manager = ProfileManager(storage)
    await profile_manager.initialize()
    
    migration_manager = ChromeMigrationManager(profile_manager)
    
    try:
        # 1. Обнаружение профилей Chrome
        logger.info("📋 Шаг 1: Обнаружение профилей Chrome")
        chrome_profiles = await migration_manager.discover_chrome_profiles()
        
        if not chrome_profiles:
            logger.warning("❌ Профили Chrome не найдены")
            logger.info("💡 Убедитесь, что Chrome установлен и есть профили с данными")
            
            # Показываем пути поиска
            paths = migration_manager.chrome_importer.chrome_data_paths
            logger.info(f"🔍 Поиск выполнялся в: {paths}")
            return
        
        logger.success(f"✅ Найдено {len(chrome_profiles)} профилей Chrome:")
        for i, profile in enumerate(chrome_profiles, 1):
            logger.info(f"  {i}. {profile['display_name']} ({profile['name']})")
            logger.info(f"     Путь: {profile['path']}")
            logger.info(f"     Куки: {'✓' if profile['has_cookies'] else '✗'}")
            logger.info(f"     Предложение: {profile['suggested_mapping']['type']}")
        
        # 2. Генерация шаблона маппинга
        logger.info("\n📝 Шаг 2: Генерация шаблона маппинга")
        template_path = await migration_manager.generate_mapping_template()
        logger.success(f"✅ Шаблон создан: {template_path}")
        
        # 3. Статус миграции
        logger.info("\n📊 Шаг 3: Текущий статус миграции")
        status = await migration_manager.get_migration_status()
        logger.info(f"Chrome профилей: {status['chrome_profiles_found']}")
        logger.info(f"Camoufox профилей всего: {status['camoufox_profiles_total']}")
        logger.info(f"Уже мигрированных: {status['migrated_profiles']}")
        
        # 4. Предварительный анализ миграции
        logger.info("\n🏃 Шаг 4: Предварительный анализ миграции")
        dry_run_results = await migration_manager.migrate_all_profiles(dry_run=True)
        
        logger.info(f"Найдено профилей для миграции: {dry_run_results['chrome_profiles_found']}")
        for result in dry_run_results['migration_results']:
            logger.info(f"  - {result['chrome_display_name']}: {result['suggested_mapping']['type']}")
        
        # 5. Интерактивная миграция с выбором профиля
        logger.info("\n❓ Шаг 5: Интерактивная миграция")
        
        if chrome_profiles:
            print("\n📋 Доступные профили Chrome для миграции:")
            for i, profile in enumerate(chrome_profiles, 1):
                status = "✓" if profile['has_cookies'] else "✗"
                print(f"  {i}. {profile['display_name']} ({profile['name']}) - Куки: {status}")
            
            print("  0. Пропустить миграцию")
            
            try:
                choice = input(f"\nВыберите профиль для миграции (1-{len(chrome_profiles)}, 0 для пропуска): ").strip()
                choice_num = int(choice)
                
                if choice_num == 0:
                    logger.info("⏭️ Миграция пропущена")
                elif 1 <= choice_num <= len(chrome_profiles):
                    selected_profile = chrome_profiles[choice_num - 1]
                    
                    logger.info(f"🔄 Начинаем миграцию профиля: {selected_profile['display_name']}")
                    logger.info(f"   Путь: {selected_profile['path']}")
                    logger.info(f"   Куки: {'Есть' if selected_profile['has_cookies'] else 'Нет'}")
                    
                    # Запрашиваем дополнительные параметры
                    print(f"\n⚙️ Настройки миграции для профиля '{selected_profile['display_name']}':")
                    
                    # Имя нового профиля
                    if selected_profile['display_name'] and selected_profile['display_name'] != selected_profile['name']:
                        default_name = f"Chrome - {selected_profile['display_name']} ({selected_profile['name']})"
                    else:
                        default_name = f"Chrome - {selected_profile['name']}"
                    new_name = input(f"Имя нового профиля [{default_name}]: ").strip()
                    if not new_name:
                        new_name = default_name
                    
                    # Группа профиля
                    default_group = "chrome_imports"
                    new_group = input(f"Группа профиля [{default_group}]: ").strip()
                    if not new_group:
                        new_group = default_group
                    
                    # Что мигрировать
                    include_cookies = input("Мигрировать куки? [Y/n]: ").strip().lower() not in ['n', 'no', 'нет']
                    include_bookmarks = input("Мигрировать закладки? [Y/n]: ").strip().lower() not in ['n', 'no', 'нет']
                    include_history = input("Мигрировать историю? [y/N]: ").strip().lower() in ['y', 'yes', 'да']
                    
                    # Создаем маппинг для миграции
                    mapping = {
                        "create_new_profile": True,
                        "new_profile_name": new_name,
                        "new_profile_group": new_group,
                        "migration_settings": {
                            "include_cookies": include_cookies,
                            "include_bookmarks": include_bookmarks,
                            "include_history": include_history
                        }
                    }
                    
                    logger.info(f"\n🚀 Запуск миграции с настройками:")
                    logger.info(f"   Новое имя: {new_name}")
                    logger.info(f"   Группа: {new_group}")
                    logger.info(f"   Куки: {'✓' if include_cookies else '✗'}")
                    logger.info(f"   Закладки: {'✓' if include_bookmarks else '✗'}")
                    logger.info(f"   История: {'✓' if include_history else '✗'}")
                    
                    result = await migration_manager.migrate_profile(selected_profile, mapping)
                    
                    if result["success"]:
                        logger.success(f"✅ Миграция успешна!")
                        logger.info(f"   Создан профиль: {result['camoufox_profile_name']} ({result['camoufox_profile_id']})")
                        logger.info(f"   Куки: {result['migration_details'].get('cookies_imported', 0)}")
                        logger.info(f"   Закладки: {result['migration_details'].get('bookmarks_imported', 0)}")
                        logger.info(f"   История: {result['migration_details'].get('history_imported', 0)}")
                    else:
                        logger.error(f"❌ Миграция неудачна:")
                        for error in result["errors"]:
                            logger.error(f"   - {error}")
                else:
                    logger.warning("⚠️ Неверный выбор профиля")
                    logger.info("⏭️ Миграция пропущена")
                    
            except ValueError:
                logger.warning("⚠️ Неверный ввод. Ожидается число.")
                logger.info("⏭️ Миграция пропущена")
            except KeyboardInterrupt:
                logger.info("\n⏭️ Миграция прервана пользователем")
        
        # 6. Показываем итоговый статус
        logger.info("\n📊 Шаг 6: Итоговый статус")
        final_status = await migration_manager.get_migration_status()
        logger.info(f"Мигрированных профилей: {final_status['migrated_profiles']}")
        
        if final_status['migrated_profile_details']:
            logger.info("Мигрированные профили:")
            for profile in final_status['migrated_profile_details']:
                logger.info(f"  - {profile['name']} (ID: {profile['id']}, Группа: {profile['group']})")
        
    except Exception as e:
        logger.error(f"❌ Ошибка миграции: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    finally:
        logger.info("🏁 Мастер миграции завершен")


async def show_excel_integration():
    """Создание примера интеграции с Excel для маппинга профилей"""
    
    logger.info("📊 Создание примера Excel интеграции")
    
    # Создаем пример Excel файла с маппингом
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Chrome Migration Mapping"
        
        # Заголовки
        headers = [
            "Chrome Profile", "Chrome Display Name", "Chrome Path",
            "Action", "Camoufox Profile ID", "New Profile Name", "New Profile Group",
            "Include Cookies", "Include Bookmarks", "Include History", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Пример данных
        example_data = [
            ["Default", "Основной", "/path/to/default", "create_new", "", "Chrome - Основной", "chrome_imports", "TRUE", "TRUE", "FALSE", "Основной профиль пользователя"],
            ["Profile 1", "Работа", "/path/to/profile1", "create_new", "", "Chrome - Работа", "work_profiles", "TRUE", "TRUE", "FALSE", "Рабочий профиль"],
            ["Profile 2", "Личный", "/path/to/profile2", "use_existing", "xcj2cs4r", "", "", "TRUE", "FALSE", "FALSE", "Перенос в существующий профиль"],
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Сохраняем файл
        excel_path = "chrome_migration_example.xlsx"
        wb.save(excel_path)
        
        logger.success(f"✅ Создан примерный Excel файл: {excel_path}")
        logger.info("📝 Этот файл можно отредактировать и использовать для массовой миграции")
        
    except ImportError:
        logger.warning("⚠️ openpyxl не установлен, пропускаем создание Excel примера")
    except Exception as e:
        logger.error(f"❌ Ошибка создания Excel файла: {e}")


def show_chrome_paths():
    """Показать пути к данным Chrome для текущей ОС"""
    from core.chrome_importer import ChromeProfileImporter
    
    importer = ChromeProfileImporter()
    
    logger.info("🗂️ Информация о путях Chrome:")
    logger.info(f"Операционная система: {importer.system}")
    logger.info(f"Путь к профилям: {importer.chrome_data_paths['profiles']}")
    
    # Проверяем существование пути
    chrome_path = Path(importer.chrome_data_paths['profiles'])
    if chrome_path.exists():
        logger.success("✅ Путь к Chrome найден")
        
        # Показываем содержимое
        try:
            subdirs = [item.name for item in chrome_path.iterdir() if item.is_dir()]
            logger.info(f"Найденные директории: {subdirs}")
        except Exception as e:
            logger.warning(f"Не удалось прочитать содержимое: {e}")
    else:
        logger.warning("⚠️ Путь к Chrome не найден")
        logger.info("💡 Возможные причины:")
        logger.info("   - Chrome не установлен")
        logger.info("   - Данные Chrome находятся в другом месте")
        logger.info("   - Нет прав доступа к директории")


async def run_bulk_migration_wizard():
    """Мастер массовой миграции профилей Chrome"""
    logger.info("🚀 Запуск мастера массовой миграции Chrome профилей")
    
    try:
        # Инициализация
        storage_manager = StorageManager("data/profiles.db")
        await storage_manager.initialize()
        
        profile_manager = ProfileManager(storage_manager, "data")
        await profile_manager.initialize()
        
        migration_manager = ChromeMigrationManager(
            profile_manager=profile_manager,
            config_path="config/chrome_migration_config.yaml"
        )
        
        # Обнаружение профилей
        chrome_profiles = await migration_manager.discover_chrome_profiles()
        
        if not chrome_profiles:
            logger.warning("❌ Профили Chrome не найдены")
            return
        
        logger.success(f"✅ Найдено {len(chrome_profiles)} профилей Chrome")
        
        print("\n📋 Доступные профили Chrome:")
        for i, profile in enumerate(chrome_profiles, 1):
            status = "✓" if profile['has_cookies'] else "✗"
            print(f"  {i}. {profile['display_name']} ({profile['name']}) - Куки: {status}")
        
        print("\n📝 Выберите профили для миграции:")
        print("   Введите номера через запятую (например: 1,3,5)")
        print("   Или 'all' для миграции всех профилей")
        print("   Или '0' для отмены")
        
        try:
            choice = input("\nВаш выбор: ").strip()
            
            if choice == "0":
                logger.info("⏭️ Массовая миграция отменена")
                return
            
            selected_profiles = []
            
            if choice.lower() == "all":
                selected_profiles = chrome_profiles
                logger.info(f"📦 Выбраны все {len(chrome_profiles)} профилей")
            else:
                # Парсим номера профилей
                profile_numbers = [int(x.strip()) for x in choice.split(",")]
                for num in profile_numbers:
                    if 1 <= num <= len(chrome_profiles):
                        selected_profiles.append(chrome_profiles[num - 1])
                    else:
                        logger.warning(f"⚠️ Неверный номер профиля: {num}")
                
                if not selected_profiles:
                    logger.warning("❌ Не выбрано ни одного валидного профиля")
                    return
                
                logger.info(f"📦 Выбрано {len(selected_profiles)} профилей:")
                for profile in selected_profiles:
                    logger.info(f"   - {profile['display_name']}")
            
            # Общие настройки миграции
            print(f"\n⚙️ Общие настройки для {len(selected_profiles)} профилей:")
            
            default_group = "chrome_bulk_import"
            new_group = input(f"Группа для всех профилей [{default_group}]: ").strip()
            if not new_group:
                new_group = default_group
            
            include_cookies = input("Мигрировать куки? [Y/n]: ").strip().lower() not in ['n', 'no', 'нет']
            include_bookmarks = input("Мигрировать закладки? [Y/n]: ").strip().lower() not in ['n', 'no', 'нет']
            include_history = input("Мигрировать историю? [y/N]: ").strip().lower() in ['y', 'yes', 'да']
            
            logger.info(f"\n🚀 Начинаем массовую миграцию {len(selected_profiles)} профилей...")
            
            successful_migrations = 0
            failed_migrations = 0
            
            for i, profile in enumerate(selected_profiles, 1):
                logger.info(f"\n📦 [{i}/{len(selected_profiles)}] Миграция: {profile['display_name']}")
                
                # Создаем уникальное имя профиля
                if profile['display_name'] and profile['display_name'] != profile['name']:
                    # Если есть отображаемое имя и оно отличается от имени директории
                    unique_name = f"Chrome - {profile['display_name']} ({profile['name']})"
                else:
                    # Если отображаемое имя отсутствует или совпадает с именем директории
                    unique_name = f"Chrome - {profile['name']}"
                
                mapping = {
                    "create_new_profile": True,
                    "new_profile_name": unique_name,
                    "new_profile_group": new_group,
                    "migration_settings": {
                        "include_cookies": include_cookies,
                        "include_bookmarks": include_bookmarks,
                        "include_history": include_history
                    }
                }
                
                try:
                    result = await migration_manager.migrate_profile(profile, mapping)
                    
                    if result["success"]:
                        successful_migrations += 1
                        logger.success(f"✅ [{i}/{len(selected_profiles)}] {profile['display_name']} - Успешно!")
                        logger.info(f"   Профиль: {result['camoufox_profile_name']} ({result['camoufox_profile_id']})")
                    else:
                        failed_migrations += 1
                        logger.error(f"❌ [{i}/{len(selected_profiles)}] {profile['display_name']} - Ошибка!")
                        for error in result["errors"]:
                            logger.error(f"   - {error}")
                
                except Exception as e:
                    failed_migrations += 1
                    logger.error(f"❌ [{i}/{len(selected_profiles)}] {profile['display_name']} - Исключение: {e}")
            
            # Итоги
            logger.info(f"\n📊 Результаты массовой миграции:")
            logger.info(f"   Всего профилей: {len(selected_profiles)}")
            logger.success(f"   Успешно: {successful_migrations}")
            if failed_migrations > 0:
                logger.error(f"   Неудачно: {failed_migrations}")
            else:
                logger.info(f"   Неудачно: {failed_migrations}")
        
        except ValueError as e:
            logger.warning(f"⚠️ Ошибка ввода: {e}")
        except KeyboardInterrupt:
            logger.info("\n⏭️ Массовая миграция прервана пользователем")
    
    except Exception as e:
        logger.error(f"❌ Ошибка массовой миграции: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    finally:
        logger.info("🏁 Мастер массовой миграции завершен")


def main():
    """Главная функция"""
    logger.info("🦊 Camoufox Chrome Migration Wizard")
    logger.info("=" * 50)
    
    # Показываем пути Chrome
    show_chrome_paths()
    
    print("\nВыберите действие:")
    print("1. Интерактивная миграция одного профиля")
    print("2. Массовая миграция нескольких профилей")
    print("3. Создать пример Excel конфигурации")
    print("4. Показать информацию о путях Chrome")
    print("0. Выход")
    
    choice = input("\nВаш выбор (1-4, 0 для выхода): ").strip()
    
    if choice == "1":
        asyncio.run(run_chrome_migration_wizard())
    elif choice == "2":
        asyncio.run(run_bulk_migration_wizard())
    elif choice == "3":
        asyncio.run(show_excel_integration())
    elif choice == "4":
        show_chrome_paths()
    elif choice == "0":
        logger.info("👋 До свидания!")
    else:
        logger.warning("⚠️ Неверный выбор")


if __name__ == "__main__":
    main() 