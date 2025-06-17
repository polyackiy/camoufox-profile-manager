"""
Модуль для экспорта и импорта профилей через Excel
"""

import io
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from loguru import logger

from .models import Profile, BrowserSettings, ProxyConfig, ProfileStatus, WebRTCMode, ProxyType
from .profile_manager import ProfileManager


class ExcelManager:
    """Менеджер для работы с Excel файлами профилей"""
    
    def __init__(self, profile_manager: ProfileManager):
        self.profile_manager = profile_manager
        
        # Определяем колонки для экспорта
        self.columns = [
            ("id", "ID профиля", "Оставьте пустым для создания нового профиля"),
            ("name", "Название профиля", "Уникальное название профиля"),
            ("group", "Группа", "Название группы профилей"),
            ("status", "Статус", "active или inactive"),
            ("os", "Операционная система", "windows, macos или linux"),
            ("screen", "Разрешение экрана", "Например: 1920x1080"),
            ("window_width", "Ширина окна", "Ширина окна браузера (800-3840)"),
            ("window_height", "Высота окна", "Высота окна браузера (600-2160)"),
            ("languages", "Языки браузера", "Через запятую: en-US, en"),
            ("timezone", "Часовой пояс", "Например: Europe/Moscow"),
            ("locale", "Локализация", "Например: ru_RU"),
            ("webrtc_mode", "Режим WebRTC", "forward, replace, real или none"),
            ("canvas_noise", "Canvas шум", "true или false"),
            ("webgl_noise", "WebGL шум", "true или false"),
            ("audio_noise", "Audio шум", "true или false"),
            ("hardware_concurrency", "Ядра процессора", "Количество ядер (1-32)"),
            ("device_memory", "Память устройства", "Память в GB (1-128)"),
            ("max_touch_points", "Точки касания", "Количество точек касания (0-10)"),
            ("proxy_type", "Тип прокси", "http, https, socks4, socks5 или пусто"),
            ("proxy_server", "Сервер прокси", "host:port или пусто"),
            ("proxy_username", "Логин прокси", "Имя пользователя или пусто"),
            ("proxy_password", "Пароль прокси", "Пароль или пусто"),
            ("geo_mode", "Режим геолокации", "auto или manual"),
            ("geo_latitude", "Широта", "Координата широты (-90 до 90)"),
            ("geo_longitude", "Долгота", "Координата долготы (-180 до 180)"),
            ("geo_accuracy", "Точность геолокации", "Точность в метрах (1-10000)"),
            ("notes", "Заметки", "Дополнительная информация о профиле")
        ]
    
    async def export_profiles_to_excel(self, profiles: Optional[List[Profile]] = None) -> bytes:
        """Экспорт профилей в Excel файл"""
        logger.info("Начинаем экспорт профилей в Excel")
        
        if profiles is None:
            profiles = await self.profile_manager.list_profiles()
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Профили браузеров"
        
        # Настраиваем стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Добавляем заголовки
        for col_idx, (field, header, comment_text) in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
            # Добавляем комментарий с пояснением
            comment = Comment(comment_text, "CamoufoxProfileManager")
            cell.comment = comment
        
        # Добавляем данные профилей
        for row_idx, profile in enumerate(profiles, 2):
            self._write_profile_row(ws, row_idx, profile)
        
        # Добавляем валидацию данных
        self._add_data_validation(ws, len(profiles) + 10)  # +10 строк для новых профилей
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в байты
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.success(f"Экспорт завершен: {len(profiles)} профилей")
        return excel_buffer.getvalue()
    
    def _write_profile_row(self, ws, row_idx: int, profile: Profile):
        """Записывает данные профиля в строку Excel"""
        browser_settings = profile.browser_settings
        proxy = profile.proxy
        
        # Подготавливаем данные
        row_data = {
            "id": profile.id,
            "name": profile.name,
            "group": profile.group or "",
            "status": profile.status.value,
            "os": browser_settings.os,
            "screen": browser_settings.screen,
            "window_width": browser_settings.window_width or "",
            "window_height": browser_settings.window_height or "",
            "languages": ", ".join(browser_settings.languages),
            "timezone": browser_settings.timezone or "",
            "locale": browser_settings.locale or "",
            "webrtc_mode": browser_settings.webrtc_mode.value,
            "canvas_noise": str(browser_settings.canvas_noise).lower(),
            "webgl_noise": str(browser_settings.webgl_noise).lower(),
            "audio_noise": str(browser_settings.audio_noise).lower(),
            "hardware_concurrency": browser_settings.hardware_concurrency or "",
            "device_memory": browser_settings.device_memory or "",
            "max_touch_points": browser_settings.max_touch_points,
            "proxy_type": proxy.type.value if proxy else "",
            "proxy_server": proxy.server if proxy else "",
            "proxy_username": proxy.username if proxy else "",
            "proxy_password": proxy.password if proxy else "",
            "geo_mode": "manual" if browser_settings.geolocation else "auto",
            "geo_latitude": browser_settings.geolocation.get("lat", "") if browser_settings.geolocation else "",
            "geo_longitude": browser_settings.geolocation.get("lon", "") if browser_settings.geolocation else "",
            "geo_accuracy": browser_settings.geolocation.get("accuracy", "") if browser_settings.geolocation else "",
            "notes": profile.notes or ""
        }
        
        # Записываем данные в строку
        for col_idx, (field, _, _) in enumerate(self.columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
