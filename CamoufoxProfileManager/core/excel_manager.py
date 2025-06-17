"""
–ú–æ–¥—É–ª—å –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ –∏ –∏–º–ø–æ—Ä—Ç–∞ –ø—Ä–æ—Ñ–∏–ª–µ–π —á–µ—Ä–µ–∑ Excel
"""

import io
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from loguru import logger

from .models import Profile, BrowserSettings, ProxyConfig, ProfileStatus, WebRTCMode, ProxyType
from .profile_manager import ProfileManager


class ExcelManager:
    """–ú–µ–Ω–µ–¥–∂–µ—Ä –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å Excel —Ñ–∞–π–ª–∞–º–∏ –ø—Ä–æ—Ñ–∏–ª–µ–π"""
    
    def __init__(self, profile_manager: ProfileManager):
        self.profile_manager = profile_manager
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞
        self.columns = [
            ("id", "ID –ø—Ä–æ—Ñ–∏–ª—è", "–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º—ã–π ID (—Ç–æ–ª—å–∫–æ –¥–ª—è —á—Ç–µ–Ω–∏—è)"),
            ("name", "–ù–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ—Ñ–∏–ª—è", "–£–Ω–∏–∫–∞–ª—å–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ—Ñ–∏–ª—è"),
            ("group", "–ì—Ä—É–ø–ø–∞", "–ù–∞–∑–≤–∞–Ω–∏–µ –≥—Ä—É–ø–ø—ã –ø—Ä–æ—Ñ–∏–ª–µ–π"),
            ("status", "–°—Ç–∞—Ç—É—Å", "active –∏–ª–∏ inactive"),
            ("os", "–û–ø–µ—Ä–∞—Ü–∏–æ–Ω–Ω–∞—è —Å–∏—Å—Ç–µ–º–∞", "windows, macos –∏–ª–∏ linux"),
            ("screen", "–†–∞–∑—Ä–µ—à–µ–Ω–∏–µ —ç–∫—Ä–∞–Ω–∞", "–ù–∞–ø—Ä–∏–º–µ—Ä: 1920x1080"),
            ("window_width", "–®–∏—Ä–∏–Ω–∞ –æ–∫–Ω–∞", "–®–∏—Ä–∏–Ω–∞ –æ–∫–Ω–∞ –±—Ä–∞—É–∑–µ—Ä–∞ (800-3840)"),
            ("window_height", "–í—ã—Å–æ—Ç–∞ –æ–∫–Ω–∞", "–í—ã—Å–æ—Ç–∞ –æ–∫–Ω–∞ –±—Ä–∞—É–∑–µ—Ä–∞ (600-2160)"),
            ("languages", "–Ø–∑—ã–∫–∏ –±—Ä–∞—É–∑–µ—Ä–∞", "–ß–µ—Ä–µ–∑ –∑–∞–ø—è—Ç—É—é: en-US, en"),
            ("timezone", "–ß–∞—Å–æ–≤–æ–π –ø–æ—è—Å", "–ù–∞–ø—Ä–∏–º–µ—Ä: Europe/Moscow"),
            ("locale", "–õ–æ–∫–∞–ª–∏–∑–∞—Ü–∏—è", "–ù–∞–ø—Ä–∏–º–µ—Ä: ru_RU"),
            ("webrtc_mode", "–†–µ–∂–∏–º WebRTC", "forward, replace, real –∏–ª–∏ none"),
            ("canvas_noise", "Canvas —à—É–º", "true –∏–ª–∏ false"),
            ("webgl_noise", "WebGL —à—É–º", "true –∏–ª–∏ false"),
            ("audio_noise", "Audio —à—É–º", "true –∏–ª–∏ false"),
            ("hardware_concurrency", "–Ø–¥—Ä–∞ –ø—Ä–æ—Ü–µ—Å—Å–æ—Ä–∞", "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —è–¥–µ—Ä (1-32)"),
            ("device_memory", "–ü–∞–º—è—Ç—å —É—Å—Ç—Ä–æ–π—Å—Ç–≤–∞", "–ü–∞–º—è—Ç—å –≤ GB (1-128)"),
            ("max_touch_points", "–¢–æ—á–∫–∏ –∫–∞—Å–∞–Ω–∏—è", "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ—á–µ–∫ –∫–∞—Å–∞–Ω–∏—è (0-10)"),
            ("proxy_type", "–¢–∏–ø –ø—Ä–æ–∫—Å–∏", "http, https, socks4, socks5 –∏–ª–∏ –ø—É—Å—Ç–æ"),
            ("proxy_server", "–°–µ—Ä–≤–µ—Ä –ø—Ä–æ–∫—Å–∏", "host:port –∏–ª–∏ –ø—É—Å—Ç–æ"),
            ("proxy_username", "–õ–æ–≥–∏–Ω –ø—Ä–æ–∫—Å–∏", "–ò–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∏–ª–∏ –ø—É—Å—Ç–æ"),
            ("proxy_password", "–ü–∞—Ä–æ–ª—å –ø—Ä–æ–∫—Å–∏", "–ü–∞—Ä–æ–ª—å –∏–ª–∏ –ø—É—Å—Ç–æ"),
            ("geo_mode", "–†–µ–∂–∏–º –≥–µ–æ–ª–æ–∫–∞—Ü–∏–∏", "auto –∏–ª–∏ manual"),
            ("geo_latitude", "–®–∏—Ä–æ—Ç–∞", "–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞ —à–∏—Ä–æ—Ç—ã (-90 –¥–æ 90)"),
            ("geo_longitude", "–î–æ–ª–≥–æ—Ç–∞", "–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞ –¥–æ–ª–≥–æ—Ç—ã (-180 –¥–æ 180)"),
            ("geo_accuracy", "–¢–æ—á–Ω–æ—Å—Ç—å –≥–µ–æ–ª–æ–∫–∞—Ü–∏–∏", "–¢–æ—á–Ω–æ—Å—Ç—å –≤ –º–µ—Ç—Ä–∞—Ö (1-10000)"),
            ("notes", "–ó–∞–º–µ—Ç–∫–∏", "–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –ø—Ä–æ—Ñ–∏–ª–µ")
        ]
    
    async def export_profiles_to_excel(self, profiles: Optional[List[Profile]] = None) -> bytes:
        """–≠–∫—Å–ø–æ—Ä—Ç –ø—Ä–æ—Ñ–∏–ª–µ–π –≤ Excel —Ñ–∞–π–ª"""
        logger.info("–ù–∞—á–∏–Ω–∞–µ–º —ç–∫—Å–ø–æ—Ä—Ç –ø—Ä–æ—Ñ–∏–ª–µ–π –≤ Excel")
        
        if profiles is None:
            profiles = await self.profile_manager.list_profiles()
        
        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∫–Ω–∏–≥—É Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "–ü—Ä–æ—Ñ–∏–ª–∏ –±—Ä–∞—É–∑–µ—Ä–æ–≤"
        
        # –ù–∞—Å—Ç—Ä–∞–∏–≤–∞–µ–º —Å—Ç–∏–ª–∏
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
        for col_idx, (field, header, comment_text) in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
            # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π —Å –ø–æ—è—Å–Ω–µ–Ω–∏–µ–º
            comment = Comment(comment_text, "CamoufoxProfileManager")
            cell.comment = comment
        
        # –î–æ–±–∞–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø—Ä–æ—Ñ–∏–ª–µ–π
        for row_idx, profile in enumerate(profiles, 2):
            self._write_profile_row(ws, row_idx, profile)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –≤–∞–ª–∏–¥–∞—Ü–∏—é –¥–∞–Ω–Ω—ã—Ö
        self._add_data_validation(ws, len(profiles) + 10)  # +10 —Å—Ç—Ä–æ–∫ –¥–ª—è –Ω–æ–≤—ã—Ö –ø—Ä–æ—Ñ–∏–ª–µ–π
        
        # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –±–∞–π—Ç—ã
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.success(f"–≠–∫—Å–ø–æ—Ä—Ç –∑–∞–≤–µ—Ä—à–µ–Ω: {len(profiles)} –ø—Ä–æ—Ñ–∏–ª–µ–π")
        return excel_buffer.getvalue()
    
    def _write_profile_row(self, ws, row_idx: int, profile: Profile):
        """–ó–∞–ø–∏—Å—ã–≤–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –ø—Ä–æ—Ñ–∏–ª—è –≤ —Å—Ç—Ä–æ–∫—É Excel"""
        browser_settings = profile.browser_settings
        proxy = profile.proxy
        
        # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ
        row_data = {
            "id": profile.id,
            "name": profile.name,
            "group": profile.group or "",
            "status": profile.status.value if hasattr(profile.status, 'value') else str(profile.status),
            "os": browser_settings.os,
            "screen": browser_settings.screen,
            "window_width": browser_settings.window_width or "",
            "window_height": browser_settings.window_height or "",
            "languages": ", ".join(browser_settings.languages),
            "timezone": browser_settings.timezone or "",
            "locale": browser_settings.locale or "",
            "webrtc_mode": browser_settings.webrtc_mode.value if hasattr(browser_settings.webrtc_mode, 'value') else str(browser_settings.webrtc_mode),
            "canvas_noise": str(browser_settings.canvas_noise).lower(),
            "webgl_noise": str(browser_settings.webgl_noise).lower(),
            "audio_noise": str(browser_settings.audio_noise).lower(),
            "hardware_concurrency": browser_settings.hardware_concurrency or "",
            "device_memory": browser_settings.device_memory or "",
            "max_touch_points": browser_settings.max_touch_points,
            "proxy_type": proxy.type.value if proxy and hasattr(proxy.type, 'value') else (str(proxy.type) if proxy else ""),
            "proxy_server": proxy.server if proxy else "",
            "proxy_username": proxy.username if proxy else "",
            "proxy_password": proxy.password if proxy else "",
            "geo_mode": "manual" if browser_settings.geolocation else "auto",
            "geo_latitude": browser_settings.geolocation.get("lat", "") if browser_settings.geolocation else "",
            "geo_longitude": browser_settings.geolocation.get("lon", "") if browser_settings.geolocation else "",
            "geo_accuracy": browser_settings.geolocation.get("accuracy", "") if browser_settings.geolocation else "",
            "notes": profile.notes or ""
        }
        
        # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ –≤ —Å—Ç—Ä–æ–∫—É
        for col_idx, (field, _, _) in enumerate(self.columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
    
    def _add_data_validation(self, ws, max_row: int):
        """–î–æ–±–∞–≤–ª—è–µ—Ç –≤–∞–ª–∏–¥–∞—Ü–∏—é –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –∫–æ–ª–æ–Ω–æ–∫"""
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è —Å—Ç–∞—Ç—É—Å–∞
        status_validation = DataValidation(
            type="list",
            formula1='"active,inactive"',
            allow_blank=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"D2:D{max_row}")
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è –û–°
        os_validation = DataValidation(
            type="list", 
            formula1='"windows,macos,linux"',
            allow_blank=True
        )
        ws.add_data_validation(os_validation)
        os_validation.add(f"E2:E{max_row}")
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è WebRTC —Ä–µ–∂–∏–º–∞
        webrtc_validation = DataValidation(
            type="list",
            formula1='"forward,replace,real,none"',
            allow_blank=True
        )
        ws.add_data_validation(webrtc_validation)
        webrtc_validation.add(f"L2:L{max_row}")
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è boolean –ø–æ–ª–µ–π
        bool_validation = DataValidation(
            type="list",
            formula1='"true,false"',
            allow_blank=True
        )
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"M2:O{max_row}")  # Canvas, WebGL, Audio noise
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è —Ç–∏–ø–∞ –ø—Ä–æ–∫—Å–∏
        proxy_validation = DataValidation(
            type="list",
            formula1='"http,https,socks4,socks5"',
            allow_blank=True
        )
        ws.add_data_validation(proxy_validation)
        proxy_validation.add(f"S2:S{max_row}")
        
        # –í–∞–ª–∏–¥–∞—Ü–∏—è –¥–ª—è —Ä–µ–∂–∏–º–∞ –≥–µ–æ–ª–æ–∫–∞—Ü–∏–∏
        geo_validation = DataValidation(
            type="list",
            formula1='"auto,manual"',
            allow_blank=True
        )
        ws.add_data_validation(geo_validation)
        geo_validation.add(f"W2:W{max_row}")
    
    async def import_profiles_from_excel(self, excel_data: bytes) -> Dict[str, Any]:
        """–ò–º–ø–æ—Ä—Ç –ø—Ä–æ—Ñ–∏–ª–µ–π –∏–∑ Excel —Ñ–∞–π–ª–∞"""
        logger.info("–ù–∞—á–∏–Ω–∞–µ–º –∏–º–ø–æ—Ä—Ç –ø—Ä–æ—Ñ–∏–ª–µ–π –∏–∑ Excel")
        
        result = {
            "success": True,
            "created_count": 0,
            "updated_count": 0,  # –í—Å–µ–≥–¥–∞ 0, —Ç–∞–∫ –∫–∞–∫ –º—ã –Ω–µ –æ–±–Ω–æ–≤–ª—è–µ–º –ø—Ä–æ—Ñ–∏–ª–∏
            "error_count": 0,
            "errors": [],
            "summary": ""
        }
        
        try:
            # –ó–∞–≥—Ä—É–∂–∞–µ–º Excel —Ñ–∞–π–ª
            excel_buffer = io.BytesIO(excel_data)
            wb = load_workbook(excel_buffer)
            ws = wb.active
            
            # –ü–æ–ª—É—á–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
            headers = {}
            for col_idx, (field, _, _) in enumerate(self.columns, 1):
                headers[field] = col_idx
            
            # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Å—Ç—Ä–æ–∫–∏
            for row_idx in range(2, ws.max_row + 1):
                try:
                    await self._process_excel_row(ws, row_idx, headers, result)
                except Exception as e:
                    error_msg = f"–°—Ç—Ä–æ–∫–∞ {row_idx}: {str(e)}"
                    result["errors"].append(error_msg)
                    result["error_count"] += 1
                    logger.error(error_msg)
            
            # –§–æ—Ä–º–∏—Ä—É–µ–º –∏—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á–µ—Ç
            result["summary"] = (
                f"–ò–º–ø–æ—Ä—Ç –∑–∞–≤–µ—Ä—à–µ–Ω:\n"
                f"‚úÖ –°–æ–∑–¥–∞–Ω–æ –ø—Ä–æ—Ñ–∏–ª–µ–π: {result['created_count']}\n"
                f"‚ùå –û—à–∏–±–æ–∫: {result['error_count']}\n"
                f"üí° –í—Å–µ –ø—Ä–æ—Ñ–∏–ª–∏ —Å–æ–∑–¥–∞—é—Ç—Å—è —Å –Ω–æ–≤—ã–º–∏ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º—ã–º–∏ ID"
            )
            
            if result["error_count"] == 0:
                logger.success(f"–ò–º–ø–æ—Ä—Ç —É—Å–ø–µ—à–µ–Ω: —Å–æ–∑–¥–∞–Ω–æ {result['created_count']} –ø—Ä–æ—Ñ–∏–ª–µ–π")
            else:
                logger.warning(f"–ò–º–ø–æ—Ä—Ç —Å –æ—à–∏–±–∫–∞–º–∏: {result['error_count']} –æ—à–∏–±–æ–∫")
                result["success"] = False
                
        except Exception as e:
            logger.error(f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –∏–º–ø–æ—Ä—Ç–∞: {e}")
            result["success"] = False
            result["errors"].append(f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞: {str(e)}")
            result["summary"] = "–ò–º–ø–æ—Ä—Ç –Ω–µ —É–¥–∞–ª—Å—è –∏–∑-–∑–∞ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–æ–π –æ—à–∏–±–∫–∏"
        
        return result
    
    async def _process_excel_row(self, ws, row_idx: int, headers: Dict[str, int], result: Dict[str, Any]):
        """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –æ–¥–Ω—É —Å—Ç—Ä–æ–∫—É Excel"""
        # –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ —Å—Ç—Ä–æ–∫–∏
        row_data = {}
        for field, col_idx in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[field] = str(cell_value).strip() if cell_value is not None else ""
        
        # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –ø—É—Å—Ç—ã–µ —Å—Ç—Ä–æ–∫–∏
        if not any(row_data.values()):
            return
        
        # –í–∞–ª–∏–¥–∏—Ä—É–µ–º –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—ã–µ –ø–æ–ª—è
        if not row_data["name"]:
            raise ValueError("–ù–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ—Ñ–∏–ª—è –æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ")
        
        # ID –∏–≥–Ω–æ—Ä–∏—Ä—É–µ—Ç—Å—è - –≤—Å–µ–≥–¥–∞ —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π –ø—Ä–æ—Ñ–∏–ª—å —Å –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º—ã–º ID
        # –≠—Ç–æ –æ–±–µ—Å–ø–µ—á–∏–≤–∞–µ—Ç —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å –∏ –ø–æ—Å–ª–µ–¥–æ–≤–∞—Ç–µ–ª—å–Ω–æ—Å—Ç—å ID
        browser_settings = self._create_browser_settings(row_data)
        proxy_config = self._create_proxy_config(row_data)
        
        await self.profile_manager.create_profile(
            name=row_data["name"],
            group=row_data["group"] or None,
            browser_settings=browser_settings,
            proxy_config=proxy_config,
            generate_fingerprint=False  # –ò—Å–ø–æ–ª—å–∑—É–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
        )
        result["created_count"] += 1
    
    def _prepare_profile_updates(self, row_data: Dict[str, str]) -> Dict[str, Any]:
        """–ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø—Ä–æ—Ñ–∏–ª—è"""
        updates = {}
        
        # –û—Å–Ω–æ–≤–Ω—ã–µ –ø–æ–ª—è –ø—Ä–æ—Ñ–∏–ª—è
        if row_data["name"]:
            updates["name"] = row_data["name"]
        if row_data["group"]:
            updates["group"] = row_data["group"]
        if row_data["status"]:
            updates["status"] = ProfileStatus(row_data["status"])
        if row_data["notes"]:
            updates["notes"] = row_data["notes"]
        
        # –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –±—Ä–∞—É–∑–µ—Ä–∞
        browser_updates = {}
        if row_data["os"]:
            browser_updates["os"] = row_data["os"]
        if row_data["screen"]:
            browser_updates["screen"] = row_data["screen"]
        if row_data["window_width"]:
            browser_updates["window_width"] = int(row_data["window_width"])
        if row_data["window_height"]:
            browser_updates["window_height"] = int(row_data["window_height"])
        if row_data["languages"]:
            browser_updates["languages"] = [lang.strip() for lang in row_data["languages"].split(",")]
        if row_data["timezone"]:
            browser_updates["timezone"] = row_data["timezone"]
        if row_data["locale"]:
            browser_updates["locale"] = row_data["locale"]
        if row_data["webrtc_mode"]:
            browser_updates["webrtc_mode"] = WebRTCMode(row_data["webrtc_mode"])
        if row_data["canvas_noise"]:
            browser_updates["canvas_noise"] = row_data["canvas_noise"].lower() == "true"
        if row_data["webgl_noise"]:
            browser_updates["webgl_noise"] = row_data["webgl_noise"].lower() == "true"
        if row_data["audio_noise"]:
            browser_updates["audio_noise"] = row_data["audio_noise"].lower() == "true"
        if row_data["hardware_concurrency"]:
            browser_updates["hardware_concurrency"] = int(row_data["hardware_concurrency"])
        if row_data["device_memory"]:
            browser_updates["device_memory"] = int(row_data["device_memory"])
        if row_data["max_touch_points"]:
            browser_updates["max_touch_points"] = int(row_data["max_touch_points"])
        
        # –ì–µ–æ–ª–æ–∫–∞—Ü–∏—è
        if row_data["geo_mode"] == "manual" and row_data["geo_latitude"] and row_data["geo_longitude"]:
            browser_updates["geolocation"] = {
                "lat": float(row_data["geo_latitude"]),
                "lon": float(row_data["geo_longitude"]),
                "accuracy": int(row_data["geo_accuracy"]) if row_data["geo_accuracy"] else 10
            }
        elif row_data["geo_mode"] == "auto":
            browser_updates["geolocation"] = None
        
        if browser_updates:
            updates["browser_settings"] = browser_updates
        
        # –ü—Ä–æ–∫—Å–∏
        if row_data["proxy_type"] and row_data["proxy_server"]:
            updates["proxy_config"] = {
                "type": row_data["proxy_type"],
                "server": row_data["proxy_server"],
                "username": row_data["proxy_username"] or None,
                "password": row_data["proxy_password"] or None
            }
        elif not row_data["proxy_type"]:
            updates["proxy_config"] = None
        
        return updates
    
    def _create_browser_settings(self, row_data: Dict[str, str]) -> BrowserSettings:
        """–°–æ–∑–¥–∞–µ—Ç –æ–±—ä–µ–∫—Ç BrowserSettings –∏–∑ –¥–∞–Ω–Ω—ã—Ö Excel"""
        # –ì–µ–æ–ª–æ–∫–∞—Ü–∏—è
        geolocation = None
        if row_data["geo_mode"] == "manual" and row_data["geo_latitude"] and row_data["geo_longitude"]:
            geolocation = {
                "lat": float(row_data["geo_latitude"]),
                "lon": float(row_data["geo_longitude"]),
                "accuracy": int(row_data["geo_accuracy"]) if row_data["geo_accuracy"] else 10
            }
        
        return BrowserSettings(
            os=row_data["os"] or "windows",
            screen=row_data["screen"] or "1920x1080",
            window_width=int(row_data["window_width"]) if row_data["window_width"] else 1280,
            window_height=int(row_data["window_height"]) if row_data["window_height"] else 720,
            languages=[lang.strip() for lang in row_data["languages"].split(",")] if row_data["languages"] else ["en-US", "en"],
            timezone=row_data["timezone"] or "UTC",
            locale=row_data["locale"] or "en_US",
            webrtc_mode=WebRTCMode(row_data["webrtc_mode"]) if row_data["webrtc_mode"] else WebRTCMode.REPLACE,
            canvas_noise=row_data["canvas_noise"].lower() == "true" if row_data["canvas_noise"] else True,
            webgl_noise=row_data["webgl_noise"].lower() == "true" if row_data["webgl_noise"] else True,
            audio_noise=row_data["audio_noise"].lower() == "true" if row_data["audio_noise"] else True,
            hardware_concurrency=int(row_data["hardware_concurrency"]) if row_data["hardware_concurrency"] else 4,
            device_memory=int(row_data["device_memory"]) if row_data["device_memory"] else 8,
            max_touch_points=int(row_data["max_touch_points"]) if row_data["max_touch_points"] else 0,
            geolocation=geolocation
        )
    
    def _create_proxy_config(self, row_data: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """–°–æ–∑–¥–∞–µ—Ç –∫–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏—é –ø—Ä–æ–∫—Å–∏ –∏–∑ –¥–∞–Ω–Ω—ã—Ö Excel"""
        if not row_data["proxy_type"] or not row_data["proxy_server"]:
            return None
        
        return {
            "type": row_data["proxy_type"],
            "server": row_data["proxy_server"],
            "username": row_data["proxy_username"] or None,
            "password": row_data["proxy_password"] or None
        }
