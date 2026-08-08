"""Tests for Excel export/import.

Import creates profiles in a loop, so a mistake here is never one profile: it is
the whole sheet. These check what a bad row costs the good ones, and what a row
is allowed to say about a profile that already exists.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook

from camoufox_pm.core.excel_manager import ExcelManager


def build_sheet(excel_manager: ExcelManager, rows: list[dict], columns=None) -> bytes:
    """Build a workbook laid out like an export, from field -> value rows.

    ``columns`` overrides the layout, to stand in for a sheet a user has edited.
    """
    workbook = Workbook()
    sheet = workbook.active
    for column, (field, header, _help) in enumerate(columns or excel_manager.columns, 1):
        sheet.cell(row=1, column=column, value=header)
        for offset, row in enumerate(rows, 2):
            sheet.cell(row=offset, column=column, value=row.get(field))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def cell(data: bytes, row: int, field: str, excel_manager: ExcelManager):
    """Read one exported cell by field name."""
    sheet = load_workbook(io.BytesIO(data)).active
    column = next(i for i, (name, _, _) in enumerate(excel_manager.columns, 1) if name == field)
    return sheet.cell(row=row, column=column).value


@pytest.mark.asyncio
async def test_export_then_import_roundtrip(profile_manager):
    await profile_manager.create_profile(name="excel-1")
    await profile_manager.create_profile(name="excel-2")

    excel_manager = ExcelManager(profile_manager)
    data = await excel_manager.export_profiles_to_excel()
    assert isinstance(data, bytes) and len(data) > 0

    result = await excel_manager.import_profiles_from_excel(data)
    # Import always creates fresh profiles, so both rows come back in.
    assert result["created_count"] == 2
    assert result["error_count"] == 0


@pytest.mark.asyncio
async def test_a_row_without_a_name_fails_alone(profile_manager):
    """One bad row in a hundred must not cost the other ninety-nine."""
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(
        excel_manager, [{"name": "first"}, {"notes": "no name here"}, {"name": "third"}]
    )

    result = await excel_manager.import_profiles_from_excel(data)

    assert result["created_count"] == 2
    assert result["error_count"] == 1
    assert result["success"] is False
    assert "Row 3" in result["errors"][0]
    assert {p.name for p in await profile_manager.list_profiles()} == {"first", "third"}


@pytest.mark.asyncio
async def test_a_row_with_an_unreadable_number_fails_alone(profile_manager):
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(
        excel_manager,
        [
            {"name": "good", "hardware_concurrency": "8"},
            {"name": "bad", "hardware_concurrency": "lots"},
        ],
    )

    result = await excel_manager.import_profiles_from_excel(data)

    assert result["created_count"] == 1
    assert result["error_count"] == 1
    assert [p.name for p in await profile_manager.list_profiles()] == ["good"]


@pytest.mark.asyncio
async def test_every_imported_row_gets_an_id_of_its_own(profile_manager):
    """A bulk import is where an ID collision hurts: profiles are stored with
    INSERT OR REPLACE, so two rows minted the same id silently became one."""
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(excel_manager, [{"name": f"bulk-{i}"} for i in range(50)])

    result = await excel_manager.import_profiles_from_excel(data)

    profiles = await profile_manager.list_profiles()
    assert result["created_count"] == 50
    assert len({p.id for p in profiles}) == 50


@pytest.mark.asyncio
async def test_a_row_carrying_an_existing_id_creates_a_profile_instead_of_replacing_it(
    profile_manager,
):
    """Re-importing an exported sheet is how people bulk-edit. The ID column is
    deliberately ignored, so the profiles it came from must survive untouched."""
    original = await profile_manager.create_profile(name="original")
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(excel_manager, [{"id": original.id, "name": "edited"}])

    await excel_manager.import_profiles_from_excel(data)

    profiles = await profile_manager.list_profiles()
    assert len(profiles) == 2
    survivor = await profile_manager.get_profile(original.id)
    assert survivor is not None and survivor.name == "original"


@pytest.mark.asyncio
async def test_a_profile_keeps_its_settings_through_an_export_and_import(profile_manager):
    """What the sheet is for: edit a fleet in Excel and get the fleet back."""
    original = await profile_manager.create_profile(
        name="detailed",
        status="inactive",
        browser_settings={
            "os": "macos",
            "screen": "2560x1440",
            "languages": ["de-DE", "de"],
            "timezone": "Europe/Berlin",
            "locale": "de_DE",
            "hardware_concurrency": 12,
            "device_memory": 16,
            "stable_canvas": True,
            "geolocation": {"lat": 52.52, "lon": 13.405, "accuracy": 25},
        },
        proxy_config={
            "type": "http",
            "server": "proxy.example:8080",
            "username": "user",
            "password": "secret",
        },
        notes="do not lose me",
    )
    excel_manager = ExcelManager(profile_manager)

    data = await excel_manager.export_profiles_to_excel()
    await excel_manager.import_profiles_from_excel(data)

    # The source profile is still there, so the import is the other one.
    imported = next(p for p in await profile_manager.list_profiles() if p.id != original.id)
    settings = imported.browser_settings
    assert imported.name == "detailed"
    assert imported.notes == "do not lose me"
    assert imported.status == "inactive"
    assert settings.os == "macos"
    assert settings.screen == "2560x1440"
    assert settings.languages == ["de-DE", "de"]
    assert settings.timezone == "Europe/Berlin"
    assert settings.locale == "de_DE"
    assert settings.hardware_concurrency == 12
    assert settings.device_memory == 16
    assert settings.stable_canvas is True
    assert settings.geolocation == {"lat": 52.52, "lon": 13.405, "accuracy": 25}
    assert imported.proxy is not None
    assert imported.proxy.server == "proxy.example:8080"
    assert imported.proxy.password == "secret"


@pytest.mark.asyncio
async def test_the_export_carries_the_proxy_password_in_clear(profile_manager):
    """It is stored encrypted, and the sheet is what makes a bulk edit possible,
    so it has to leave in clear. Anyone handing this file on should know."""
    await profile_manager.create_profile(
        name="p",
        proxy_config={"type": "http", "server": "h:8080", "username": "u", "password": "secret"},
    )
    excel_manager = ExcelManager(profile_manager)

    data = await excel_manager.export_profiles_to_excel()

    assert cell(data, 2, "proxy_password", excel_manager) == "secret"


@pytest.mark.asyncio
async def test_blank_rows_are_skipped_rather_than_counted(profile_manager):
    """Spreadsheets carry trailing blank rows; each one used to be an error."""
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(excel_manager, [{"name": "only"}, {}, {}, {}])

    result = await excel_manager.import_profiles_from_excel(data)

    assert result["created_count"] == 1
    assert result["error_count"] == 0
    assert result["success"] is True


@pytest.mark.asyncio
async def test_manual_geolocation_without_coordinates_falls_back_to_the_proxy(profile_manager):
    """Half-filled coordinates would place the profile at (0, 0), off West Africa."""
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(excel_manager, [{"name": "p", "geo_mode": "manual"}])

    await excel_manager.import_profiles_from_excel(data)

    imported = (await profile_manager.list_profiles())[0]
    assert imported.browser_settings.geolocation is None


@pytest.mark.asyncio
async def test_an_unknown_status_fails_the_row_instead_of_being_ignored(profile_manager):
    """Silently importing it as active would put a parked profile back to work."""
    excel_manager = ExcelManager(profile_manager)
    data = build_sheet(excel_manager, [{"name": "p", "status": "paused"}])

    result = await excel_manager.import_profiles_from_excel(data)

    assert result["created_count"] == 0
    assert result["error_count"] == 1


@pytest.mark.asyncio
async def test_a_moved_column_is_read_from_where_it_now_is(profile_manager):
    """Regression: columns were read by position, so a sheet the user rearranged
    imported cleanly with the values in the wrong fields. Moving Locale ahead of
    Timezone gave every profile a timezone of "de_DE" and a locale of
    "Europe/Berlin", and nothing objected — most of these fields are free text.
    """
    excel_manager = ExcelManager(profile_manager)
    rearranged = list(excel_manager.columns)
    timezone_at = next(i for i, c in enumerate(rearranged) if c[0] == "timezone")
    locale_at = next(i for i, c in enumerate(rearranged) if c[0] == "locale")
    rearranged[timezone_at], rearranged[locale_at] = (
        rearranged[locale_at],
        rearranged[timezone_at],
    )
    data = build_sheet(
        excel_manager,
        [{"name": "p", "timezone": "Europe/Berlin", "locale": "de_DE"}],
        columns=rearranged,
    )

    await excel_manager.import_profiles_from_excel(data)

    imported = (await profile_manager.list_profiles())[0]
    assert imported.browser_settings.timezone == "Europe/Berlin"
    assert imported.browser_settings.locale == "de_DE"


@pytest.mark.asyncio
async def test_a_deleted_column_leaves_its_field_at_the_default(profile_manager):
    """The ID column is marked read-only, so deleting it before editing is the
    obvious thing to do. It must not shift every column after it."""
    excel_manager = ExcelManager(profile_manager)
    without_id = [c for c in excel_manager.columns if c[0] != "id"]
    data = build_sheet(
        excel_manager,
        [{"name": "p", "os": "macos", "notes": "kept"}],
        columns=without_id,
    )

    result = await excel_manager.import_profiles_from_excel(data)

    assert result["error_count"] == 0
    imported = (await profile_manager.list_profiles())[0]
    assert imported.name == "p"
    assert imported.browser_settings.os == "macos"
    assert imported.notes == "kept"


@pytest.mark.asyncio
async def test_a_sheet_with_no_recognisable_header_row_is_refused(profile_manager):
    """Importing it row by row would create a fleet of mangled profiles."""
    excel_manager = ExcelManager(profile_manager)
    workbook = Workbook()
    sheet = workbook.active
    for column, value in enumerate(["Column A", "Column B"], 1):
        sheet.cell(row=1, column=column, value=value)
        sheet.cell(row=2, column=column, value="something")
    buffer = io.BytesIO()
    workbook.save(buffer)

    result = await excel_manager.import_profiles_from_excel(buffer.getvalue())

    assert result["success"] is False
    assert result["created_count"] == 0
    assert await profile_manager.list_profiles() == []
    assert "Profile name" in result["errors"][0]


@pytest.mark.asyncio
async def test_a_file_that_is_not_a_workbook_is_reported_not_raised(profile_manager):
    result = await ExcelManager(profile_manager).import_profiles_from_excel(b"not a workbook")

    assert result["success"] is False
    assert result["created_count"] == 0
    assert result["errors"]
